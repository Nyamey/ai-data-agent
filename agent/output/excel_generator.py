# agent/output/excel_generator.py : génération de classeurs Excel
import io

import openpyxl
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from agent.tools.spreadsheet_safety import neutralize_formulas, unique_sheet_title


class ExcelGenerator:
    """
    Génère un classeur Excel avec analyses et graphiques.

    Onglets générés :
    1. Métrique construite (ex. rétention hebdomadaire)
    2. Analyse des facteurs, un onglet par dimension testée
    3. Validation
    """

    def __init__(self):
        self.wb = openpyxl.Workbook()
        # Style des en-têtes
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="2F5496", end_color="2F5496")
        self._used_titles = set()

    def _add_sheet_with_data(self, title: str, df: pd.DataFrame, chart_type: str = None):
        """Ajoute une feuille avec des données et optionnellement un graphique."""
        title = unique_sheet_title(title, self._used_titles)
        ws = self.wb.create_sheet(title=title)

        # Ajouter les données
        for row in dataframe_to_rows(neutralize_formulas(df), index=False, header=True):
            ws.append(row)
        
        # Styliser l'en-tête
        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-largeur des colonnes
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 30)
        
        # Ajouter un graphique si demandé
        if chart_type == "line" and ws.max_row > 1:
            chart = LineChart()
            chart.title = title
            data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row, max_col=ws.max_column)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, f"A{ws.max_row + 3}")
        
        elif chart_type == "bar" and ws.max_row > 1:
            chart = BarChart()
            chart.title = title
            data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row, max_col=ws.max_column)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, f"A{ws.max_row + 3}")
    
    def add_weekly_retention(self, df: pd.DataFrame, title: str = "Rétention Hebdo"):
        """Ajoute l'onglet de métrique construite (rétention hebdomadaire ou équivalent générique)."""
        self._add_sheet_with_data(title, df, chart_type="line")

    def add_driver_analysis(self, df: pd.DataFrame, title: str = "Analyse Facteurs"):
        """Ajoute un onglet d'analyse de facteur (appelable plusieurs fois, un par dimension)."""
        self._add_sheet_with_data(title, df, chart_type="bar")
    
    def add_cleaned_data(self, df: pd.DataFrame, title: str = "Données nettoyées"):
        """Ajoute la donnée nettoyée elle-même, suivie d'une feuille de
        statistiques descriptives (df.describe()), pensée pour un export
        autonome des données de l'analyse, séparé du rapport de résultats
        (métrique construite, facteurs, validation, recommandations).

        df.describe() décrit les colonnes numériques par défaut, ou bascule
        automatiquement sur des statistiques catégorielles (count/unique/
        top/freq) s'il n'y a aucune colonne numérique, comportement natif
        de pandas, pas de cas particulier à gérer ici.
        """
        self._add_sheet_with_data(title, df)

        stats_title = unique_sheet_title(f"Stats - {title}", self._used_titles)
        ws = self.wb.create_sheet(title=stats_title)
        # openpyxl.dataframe_to_rows(index=True, header=True) insère toujours
        # une ligne fantôme juste après l'en-tête (le nom de l'index, sur sa
        # propre ligne), un artefact de sa mise en forme, pas une donnée à
        # garder.
        for i, row in enumerate(dataframe_to_rows(neutralize_formulas(df.describe()), index=True, header=True)):
            if i != 1:
                ws.append(row)
        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill

    def add_validation_checks(self, checks: dict):
        """Ajoute l'onglet validation."""
        ws = self.wb.create_sheet(title=unique_sheet_title("Validation", self._used_titles))
        ws.append(["Contrôle", "Résultat", "Statut"])
        
        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        for key, value in checks.items():
            status = "OK" if value.get("passed") else "ÉCHEC"
            ws.append([key, str(value.get("result", "")), status])
    
    def _drop_default_sheet(self):
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

    def save(self, path: str) -> str:
        """Sauvegarde le classeur sur disque et retourne le chemin."""
        self._drop_default_sheet()
        self.wb.save(path)
        return path

    def to_bytes(self) -> bytes:
        """Sérialise le classeur en mémoire, pour un téléchargement Streamlit
        direct (`st.download_button`), sans passer par un fichier temporaire."""
        self._drop_default_sheet()
        buf = io.BytesIO()
        self.wb.save(buf)
        return buf.getvalue()
