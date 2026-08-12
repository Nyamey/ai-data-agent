# tests/test_excel_generator.py — Génération de classeurs Excel
import openpyxl
import pandas as pd

from agent.output.excel_generator import ExcelGenerator


def test_add_weekly_retention_creates_sheet_with_data(tmp_path):
    gen = ExcelGenerator()
    df = pd.DataFrame({"semaine": ["2024-01-01", "2024-01-08"], "valeur": [10, 12]})
    gen.add_weekly_retention(df, title="Rétention")
    path = str(tmp_path / "out.xlsx")
    gen.save(path)

    wb = openpyxl.load_workbook(path)
    assert "Rétention" in wb.sheetnames
    assert "Sheet" not in wb.sheetnames


def test_sheet_titles_are_deduplicated(tmp_path):
    gen = ExcelGenerator()
    df = pd.DataFrame({"a": [1], "b": [2]})
    gen.add_driver_analysis(df, title="Facteur - plateforme")
    gen.add_driver_analysis(df, title="Facteur - plateforme")
    path = str(tmp_path / "out.xlsx")
    gen.save(path)

    wb = openpyxl.load_workbook(path)
    sheets = [s for s in wb.sheetnames if s != "Sheet"]
    assert len(sheets) == 2
    assert len(set(sheets)) == 2


def test_sheet_title_truncated_to_excel_limit(tmp_path):
    gen = ExcelGenerator()
    df = pd.DataFrame({"a": [1]})
    long_title = "Facteur - " + "x" * 50
    gen.add_driver_analysis(df, title=long_title)
    path = str(tmp_path / "out.xlsx")
    gen.save(path)

    wb = openpyxl.load_workbook(path)
    sheets = [s for s in wb.sheetnames if s != "Sheet"]
    assert all(len(s) <= 31 for s in sheets)


def test_add_validation_checks_reports_pass_fail(tmp_path):
    gen = ExcelGenerator()
    gen.add_validation_checks({
        "doublons": {"result": 0, "passed": True},
        "plage_dates": {"result": "erreur", "passed": False},
    })
    path = str(tmp_path / "out.xlsx")
    gen.save(path)

    wb = openpyxl.load_workbook(path)
    ws = wb["Validation"]
    rows = list(ws.iter_rows(values_only=True))
    statuses = {r[0]: r[2] for r in rows[1:]}
    assert statuses["doublons"] == "OK"
    assert statuses["plage_dates"] == "ÉCHEC"


def test_add_cleaned_data_creates_data_and_stats_sheets(tmp_path):
    gen = ExcelGenerator()
    df = pd.DataFrame({"montant": [10, 20, 30], "plateforme": ["mobile", "web", "mobile"]})
    gen.add_cleaned_data(df, title="Données nettoyées")
    path = str(tmp_path / "out.xlsx")
    gen.save(path)

    wb = openpyxl.load_workbook(path)
    assert "Données nettoyées" in wb.sheetnames
    assert "Stats - Données nettoyées" in wb.sheetnames

    data_rows = list(wb["Données nettoyées"].iter_rows(values_only=True))
    assert data_rows[0] == ("montant", "plateforme")
    assert len(data_rows) == 4  # en-tête + 3 lignes

    stats_rows = list(wb["Stats - Données nettoyées"].iter_rows(values_only=True))
    # Pas de ligne fantôme entre l'en-tête et la première statistique --
    # voir le commentaire dans excel_generator.py sur ce comportement de
    # dataframe_to_rows(index=True, header=True).
    assert stats_rows[1][0] == "count"
    assert stats_rows[1][1] == 3.0


def test_add_cleaned_data_falls_back_to_categorical_stats_without_numeric_columns(tmp_path):
    gen = ExcelGenerator()
    df = pd.DataFrame({"ville": ["Paris", "Lyon", "Paris"]})
    gen.add_cleaned_data(df, title="Texte seul")
    path = str(tmp_path / "out.xlsx")
    gen.save(path)

    wb = openpyxl.load_workbook(path)
    stats_rows = list(wb["Stats - Texte seul"].iter_rows(values_only=True))
    labels = [r[0] for r in stats_rows[1:]]
    assert "count" in labels
    assert "top" in labels  # statistique catégorielle, pandas bascule dessus sans colonne numérique


def test_add_cleaned_data_neutralizes_formula_injection_in_data_and_headers(tmp_path):
    gen = ExcelGenerator()
    df = pd.DataFrame({'=HYPERLINK("http://evil")': ["+1+1", "safe"]})
    gen.add_cleaned_data(df, title="Données")
    path = str(tmp_path / "out.xlsx")
    gen.save(path)

    wb = openpyxl.load_workbook(path)
    ws = wb["Données"]
    header = list(ws.iter_rows(max_row=1, values_only=True))[0]
    assert header[0].startswith("'=")
    values = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert values[0].startswith("'+")


def test_to_bytes_produces_a_valid_workbook_without_writing_to_disk():
    import io

    gen = ExcelGenerator()
    gen.add_cleaned_data(pd.DataFrame({"a": [1, 2]}), title="Données")
    data = gen.to_bytes()

    assert data[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Sheet" not in wb.sheetnames
    assert "Données" in wb.sheetnames


def test_driver_analysis_neutralizes_formula_injection(tmp_path):
    # Régression (CWE-1236) : une valeur de catégorie du CSV téléversé par
    # l'utilisateur (ex. une valeur de "plateforme") atterrit telle quelle
    # dans un onglet -- sans neutralisation, Excel l'interpréterait comme une
    # formule à l'ouverture.
    gen = ExcelGenerator()
    df = pd.DataFrame({
        "plateforme": ['=HYPERLINK("http://evil")', "+1+1", "mobile"],
        "nb_entites": [3, 2, 40],
    })
    gen.add_driver_analysis(df, title="Facteur - plateforme")
    path = str(tmp_path / "out.xlsx")
    gen.save(path)

    wb = openpyxl.load_workbook(path)
    ws = wb["Facteur - plateforme"]
    values = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert values[0].startswith("'=")
    assert values[1].startswith("'+")
    assert values[2] == "mobile"
